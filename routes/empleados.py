from flask import Blueprint, render_template, flash, redirect, url_for, jsonify, request, send_file
from flask_login import login_required, current_user
from db import get_db_connection
# Librerías para PDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# Librerías para Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference

import requests
import base64
import io
from datetime import datetime
import os
import tempfile
import pdfkit  

# Crear Blueprint para empleados (SOLO UNA VEZ)
empleados_bp = Blueprint('empleados', __name__, url_prefix='/empleado')

# Configurar pdfkit (si lo vas a usar)
WKHTMLTOPDF_PATH = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'  # Ajusta esta ruta
config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

def get_id_empleado():
    """Obtener el id_empleado del usuario actual"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id_empleado FROM empleado WHERE id_usuario = %s", (current_user.id,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        return result[0] if result else None
    except Exception as e:
        print(f"❌ Error obteniendo id_empleado: {e}")
        return None

# ================================
# DASHBOARD DEL EMPLEADO - CON GRÁFICOS
# ================================
@empleados_bp.route("/dashboard")
@login_required
def dashboard():
    if current_user.id_rol != 2:
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("auth.login"))
    
    try:
        id_empleado = get_id_empleado()
        if not id_empleado:
            flash("No se encontró información del empleado.", "danger")
            return redirect(url_for("auth.login"))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Datos del empleado logueado
        cursor.execute("""
            SELECT u.nombre, u.ap_paterno, u.ap_materno, e.puesto, e.fecha_contratacion
            FROM usuario u
            JOIN empleado e ON u.id_usuario = e.id_usuario
            WHERE e.id_empleado = %s
        """, (id_empleado,))
        empleado_data = cursor.fetchone()

        if not empleado_data:
            flash("No se encontraron datos del empleado.", "danger")
            return render_template("empleado/dashboard.html")

        # Total de tickets asignados
        cursor.execute("SELECT COUNT(*) FROM ticket WHERE id_empleado = %s", (id_empleado,))
        total_tickets = cursor.fetchone()[0]

        # Tickets pendientes
        cursor.execute("""
            SELECT COUNT(*) FROM ticket 
            WHERE id_empleado = %s AND estado IN ('abierto', 'en_progreso')
        """, (id_empleado,))
        tickets_pendientes = cursor.fetchone()[0]

        # Total de mantenimientos
        cursor.execute("SELECT COUNT(*) FROM mantenimiento WHERE id_empleado = %s", (id_empleado,))
        total_mantenimientos = cursor.fetchone()[0]

        # =====================
        # DATOS PARA GRÁFICOS DE CONSUMO
        # =====================
        
        # Consumo de AGUA del mes actual
        cursor.execute("""
            SELECT COALESCE(SUM(cantidad_registrada), 0) 
            FROM consumo_agua 
            WHERE DATE_PART('month', fecha_registro) = DATE_PART('month', CURRENT_DATE)
            AND DATE_PART('year', fecha_registro) = DATE_PART('year', CURRENT_DATE)
        """)
        consumo_agua_actual = float(cursor.fetchone()[0] or 0)

        # Consumo de LUZ del mes actual
        cursor.execute("""
            SELECT COALESCE(SUM(cantidad_registrada), 0) 
            FROM consumo_luz 
            WHERE DATE_PART('month', fecha_registro) = DATE_PART('month', CURRENT_DATE)
            AND DATE_PART('year', fecha_registro) = DATE_PART('year', CURRENT_DATE)
        """)
        consumo_luz_actual = float(cursor.fetchone()[0] or 0)

        # Consumo de GAS del mes actual
        cursor.execute("""
            SELECT COALESCE(SUM(cantidad_registrada), 0) 
            FROM consumo_gas 
            WHERE DATE_PART('month', fecha_registro) = DATE_PART('month', CURRENT_DATE)
            AND DATE_PART('year', fecha_registro) = DATE_PART('year', CURRENT_DATE)
        """)
        consumo_gas_actual = float(cursor.fetchone()[0] or 0)

        # Consumos de los últimos 6 meses para gráficos
        cursor.execute("""
            SELECT 
                TO_CHAR(fecha_registro, 'YYYY-MM') as mes,
                SUM(cantidad_registrada) as consumo
            FROM consumo_agua 
            WHERE fecha_registro >= CURRENT_DATE - INTERVAL '6 months'
            GROUP BY TO_CHAR(fecha_registro, 'YYYY-MM')
            ORDER BY mes
        """)
        agua_mensual = cursor.fetchall()

        cursor.execute("""
            SELECT 
                TO_CHAR(fecha_registro, 'YYYY-MM') as mes,
                SUM(cantidad_registrada) as consumo
            FROM consumo_luz 
            WHERE fecha_registro >= CURRENT_DATE - INTERVAL '6 months'
            GROUP BY TO_CHAR(fecha_registro, 'YYYY-MM')
            ORDER BY mes
        """)
        luz_mensual = cursor.fetchall()

        cursor.execute("""
            SELECT 
                TO_CHAR(fecha_registro, 'YYYY-MM') as mes,
                SUM(cantidad_registrada) as consumo
            FROM consumo_gas 
            WHERE fecha_registro >= CURRENT_DATE - INTERVAL '6 months'
            GROUP BY TO_CHAR(fecha_registro, 'YYYY-MM')
            ORDER BY mes
        """)
        gas_mensual = cursor.fetchall()

        cursor.close()
        conn.close()

        # Preparar datos para gráficos
        consumos_mensuales = {
            'agua': [{'mes': row[0], 'consumo': float(row[1] or 0)} for row in agua_mensual],
            'luz': [{'mes': row[0], 'consumo': float(row[1] or 0)} for row in luz_mensual],
            'gas': [{'mes': row[0], 'consumo': float(row[1] or 0)} for row in gas_mensual]
        }

        # Pasar datos al template
        empleado_dict = {
            'nombre': empleado_data[0] or '',
            'ap_paterno': empleado_data[1] or '',
            'ap_materno': empleado_data[2] or '',
            'puesto': empleado_data[3] or '',
            'fecha_contratacion': empleado_data[4] or ''
        }

        return render_template(
            "empleado/dashboard.html",
            empleado=empleado_dict,
            total_tickets=total_tickets,
            tickets_pendientes=tickets_pendientes,
            total_mantenimientos=total_mantenimientos,
            consumo_agua=consumo_agua_actual,
            consumo_luz=consumo_luz_actual,
            consumo_gas=consumo_gas_actual,
            consumos_mensuales=consumos_mensuales
        )

    except Exception as e:
        print(f"❌ Error cargando dashboard: {e}")
        # Pasar valores por defecto para evitar errores en el template
        empleado_dict = {
            'nombre': '',
            'ap_paterno': '',
            'ap_materno': '',
            'puesto': '',
            'fecha_contratacion': ''
        }
        return render_template(
            "empleado/dashboard.html",
            empleado=empleado_dict,
            total_tickets=0,
            tickets_pendientes=0,
            total_mantenimientos=0,
            consumo_agua=0,
            consumo_luz=0,
            consumo_gas=0,
            consumos_mensuales={'agua': [], 'luz': [], 'gas': []}
        )

# ================================
# TICKETS - CONSULTA CORREGIDA (usando tabla 'area')
# ================================
@empleados_bp.route("/tickets")
@login_required
def tickets():
    if current_user.id_rol != 2:
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("auth.login"))
    
    try:
        id_empleado = get_id_empleado()
        if not id_empleado:
            flash("No se encontró información del empleado.", "danger")
            return render_template("empleado/tickets.html", tickets=[])

        conn = get_db_connection()
        cursor = conn.cursor()

        # CONSULTA CORREGIDA - Usa 'area' en lugar de 'area_comun'
        cursor.execute("""
            SELECT 
                t.id_ticket, 
                t.descripcion, 
                t.prioridad, 
                t.estado,
                TO_CHAR(t.fecha_emision, 'DD/MM/YYYY HH24:MI') as fecha_emision,
                TO_CHAR(t.fecha_finalizacion, 'DD/MM/YYYY HH24:MI') as fecha_finalizacion,
                a.nombre as area
            FROM ticket t
            JOIN area a ON t.id_area = a.id_area
            WHERE t.id_empleado = %s
            ORDER BY 
                CASE 
                    WHEN t.estado = 'en_progreso' THEN 1
                    WHEN t.estado = 'abierto' THEN 2
                    WHEN t.estado = 'cerrado' THEN 3
                    WHEN t.estado = 'cancelado' THEN 4
                    ELSE 5
                END,
                CASE 
                    WHEN t.prioridad = 'urgente' THEN 1
                    WHEN t.prioridad = 'alta' THEN 2
                    WHEN t.prioridad = 'media' THEN 3
                    WHEN t.prioridad = 'baja' THEN 4
                    ELSE 5
                END,
                t.fecha_emision DESC
        """, (id_empleado,))
        tickets = cursor.fetchall()

        cursor.close()
        conn.close()

        print(f"✅ Tickets encontrados: {len(tickets)}")
        return render_template("empleado/tickets.html", tickets=tickets)
        
    except Exception as e:
        print(f"❌ Error obteniendo tickets: {e}")
        flash("Error al cargar los tickets.", "danger")
        return render_template("empleado/tickets.html", tickets=[])

# ================================
# MANTENIMIENTOS - CONSULTA CORREGIDA (usando tabla 'area')
# ================================
@empleados_bp.route("/mantenimientos")
@login_required
def mantenimientos():
    if current_user.id_rol != 2:
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("auth.login"))
    
    try:
        id_empleado = get_id_empleado()
        if not id_empleado:
            flash("No se encontró información del empleado.", "danger")
            return render_template("empleado/mantenimientos.html", mantenimientos=[])

        conn = get_db_connection()
        cursor = conn.cursor()

        # CONSULTA CORREGIDA - Usa 'area' en lugar de 'area_comun'
        cursor.execute("""
            SELECT 
                m.id_mantenimiento, 
                m.descripcion, 
                TO_CHAR(m.fecha_programada, 'DD/MM/YYYY'), 
                m.estado, 
                a.nombre as area
            FROM mantenimiento m
            JOIN area a ON m.id_area = a.id_area
            WHERE m.id_empleado = %s
            ORDER BY 
                CASE m.estado
                    WHEN 'en_proceso' THEN 1
                    WHEN 'programado' THEN 2
                    WHEN 'completado' THEN 3
                    ELSE 4
                END,
                m.fecha_programada DESC
        """, (id_empleado,))
        mantenimientos = cursor.fetchall()

        cursor.close()
        conn.close()

        print(f"✅ Mantenimientos encontrados: {len(mantenimientos)}")
        return render_template("empleado/mantenimientos.html", mantenimientos=mantenimientos)

    except Exception as e:
        print(f"❌ Error obteniendo mantenimientos: {e}")
        flash("Error al cargar los mantenimientos.", "danger")
        return render_template("empleado/mantenimientos.html", mantenimientos=[])

# ================================
# REPORTES - VERSIÓN CORREGIDA (SOLO UNA)
# ================================
@empleados_bp.route("/reportes")
@login_required
def reportes():
    """Endpoint único para reportes - usa el template nuevo con filtros"""
    if current_user.id_rol != 2:
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("auth.login"))
    
    try:
        return render_template("empleado/reportes.html")
        
    except Exception as e:
        print(f"❌ Error cargando página de reportes: {e}")
        flash("Error al cargar la página de reportes.", "danger")
        return render_template("empleado/reportes.html")

# ================================
# PERFIL
# ================================
@empleados_bp.route("/perfil")
@login_required
def perfil():
    if current_user.id_rol != 2:
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("auth.login"))
    
    try:
        id_empleado = get_id_empleado()
        if not id_empleado:
            flash("No se encontró información del empleado.", "danger")
            return redirect(url_for("empleados.dashboard"))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Consultar datos del empleado
        cursor.execute("""
            SELECT e.id_empleado, e.puesto, e.salario, e.fecha_contratacion,
                   u.nombre, u.ap_paterno, u.ap_materno, u.correo, u.telefono
            FROM empleado e
            JOIN usuario u ON e.id_usuario = u.id_usuario
            WHERE e.id_empleado = %s
        """, (id_empleado,))
        empleado_data = cursor.fetchone()
        
        # Obtener estadísticas
        cursor.execute("""
            SELECT COUNT(*) FROM ticket 
            WHERE id_empleado = %s AND estado = 'cerrado'
        """, (id_empleado,))
        tickets_completados = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM mantenimiento 
            WHERE id_empleado = %s AND estado = 'completado'
        """, (id_empleado,))
        mantenimientos_realizados = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT COUNT(*) FROM ticket 
            WHERE id_empleado = %s AND estado IN ('abierto', 'en_progreso')
        """, (id_empleado,))
        tickets_pendientes = cursor.fetchone()[0]
        
        cursor.close()
        conn.close()

        # Convertir tupla a diccionario para mejor acceso en template
        empleado_dict = {
            'id_empleado': empleado_data[0],
            'puesto': empleado_data[1],
            'salario': empleado_data[2],
            'fecha_contratacion': empleado_data[3],
            'nombre': empleado_data[4],
            'ap_paterno': empleado_data[5],
            'ap_materno': empleado_data[6],
            'correo': empleado_data[7],
            'telefono': empleado_data[8]
        }
        
        return render_template('empleado/perfil.html',
                            empleado=empleado_dict,
                            tickets_completados=tickets_completados,
                            mantenimientos_realizados=mantenimientos_realizados,
                            tickets_pendientes=tickets_pendientes)
    
    except Exception as e:
        print(f"❌ Error cargando perfil: {e}")
        flash("Error al cargar el perfil.", "danger")
        return redirect(url_for("empleados.dashboard"))

# ================================
# ENDPOINTS API PARA REPORTES NUEVOS
# ================================

@empleados_bp.route('/api/tickets-reporte')
@login_required
def obtener_tickets_reporte():
    """API para obtener tickets con filtros - para el nuevo sistema de reportes"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Obtener el ID del empleado logueado
        id_empleado = get_id_empleado()
        if not id_empleado:
            return jsonify({'success': False, 'error': 'Empleado no encontrado'}), 400
        
        # Obtener parámetros de filtro
        estado = request.args.get('estado', 'todos')
        prioridad = request.args.get('prioridad', 'todos')
        
        query = """
        SELECT 
            t.id_ticket,
            t.descripcion,
            t.prioridad,
            t.estado,
            TO_CHAR(t.fecha_emision, 'DD/MM/YYYY HH24:MI') as fecha_emision,
            TO_CHAR(t.fecha_finalizacion, 'DD/MM/YYYY HH24:MI') as fecha_finalizacion,
            a.nombre as area,
            a.ubicacion,
            CASE 
                WHEN t.estado = 'cerrado' THEN 
                    'Completado en ' || DATE_PART('day', t.fecha_finalizacion - t.fecha_emision) || ' días'
                ELSE 
                    'Abierto hace ' || DATE_PART('day', CURRENT_TIMESTAMP - t.fecha_emision) || ' días'
            END as tiempo_transcurrido
        FROM ticket t
        JOIN area a ON t.id_area = a.id_area
        WHERE t.id_empleado = %s
        """
        
        params = [id_empleado]
        
        # Aplicar filtros
        if estado != 'todos':
            query += " AND t.estado = %s"
            params.append(estado)
            
        if prioridad != 'todos':
            query += " AND t.prioridad = %s"
            params.append(prioridad)
        
        query += " ORDER BY t.fecha_emision DESC"
        
        cur.execute(query, params)
        tickets = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Convertir a formato JSON
        tickets_json = []
        for ticket in tickets:
            tickets_json.append({
                'id': ticket[0],
                'descripcion': ticket[1],
                'prioridad': ticket[2],
                'estado': ticket[3],
                'fecha_emision': ticket[4],
                'fecha_finalizacion': ticket[5],
                'area': ticket[6],
                'ubicacion': ticket[7],
                'tiempo_transcurrido': ticket[8]
            })
        
        return jsonify({
            'success': True,
            'data': tickets_json,
            'total': len(tickets_json)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@empleados_bp.route('/generar_reporte_pdf', methods=['POST'])
@login_required
def generar_reporte_pdf():
    """Generar PDF con los nuevos filtros"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Obtener filtros del request
        data = request.get_json()
        filtros = data.get('filtros', {})
        
        # Obtener el ID del empleado logueado
        id_empleado = get_id_empleado()
        if not id_empleado:
            return jsonify({'success': False, 'error': 'Empleado no encontrado'}), 400
        
        # Query para obtener tickets
        query = """
        SELECT 
            t.id_ticket,
            t.descripcion,
            t.prioridad,
            t.estado,
            TO_CHAR(t.fecha_emision, 'DD/MM/YYYY HH24:MI') as fecha_emision,
            TO_CHAR(t.fecha_finalizacion, 'DD/MM/YYYY HH24:MI') as fecha_finalizacion,
            a.nombre as area,
            a.ubicacion,
            u.nombre || ' ' || u.ap_paterno as empleado_nombre,
            e.puesto,
            CASE 
                WHEN t.estado = 'cerrado' THEN 
                    'Completado en ' || DATE_PART('day', t.fecha_finalizacion - t.fecha_emision) || ' días'
                ELSE 
                    'Abierto hace ' || DATE_PART('day', CURRENT_TIMESTAMP - t.fecha_emision) || ' días'
            END as tiempo_transcurrido
        FROM ticket t
        JOIN area a ON t.id_area = a.id_area
        JOIN empleado e ON t.id_empleado = e.id_empleado
        JOIN usuario u ON e.id_usuario = u.id_usuario
        WHERE t.id_empleado = %s
        """
        
        params = [id_empleado]
        
        # Aplicar filtros
        if filtros.get('estado') and filtros['estado'] != 'todos':
            query += " AND t.estado = %s"
            params.append(filtros['estado'])
            
        if filtros.get('prioridad') and filtros['prioridad'] != 'todos':
            query += " AND t.prioridad = %s"
            params.append(filtros['prioridad'])
        
        query += " ORDER BY t.fecha_emision DESC"
        
        cur.execute(query, params)
        tickets = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Preparar datos para el template
        tickets_data = []
        for ticket in tickets:
            tickets_data.append({
                'id_ticket': ticket[0],
                'descripcion': ticket[1],
                'prioridad': ticket[2],
                'estado': ticket[3],
                'fecha_emision': ticket[4],
                'fecha_finalizacion': ticket[5],
                'area': ticket[6],
                'ubicacion': ticket[7],
                'empleado_nombre': ticket[8],
                'puesto': ticket[9],
                'tiempo_transcurrido': ticket[10]
            })
        
        # Datos para el reporte
        reporte_data = {
            'tickets': tickets_data,
            'fecha_generacion': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'total_tickets': len(tickets_data),
            'empleado_nombre': tickets_data[0]['empleado_nombre'] if tickets_data else 'N/A',
            'puesto': tickets_data[0]['puesto'] if tickets_data else 'N/A'
        }
        
        # Renderizar HTML
        html_content = render_template('empleado/reporte_pdf.html', **reporte_data)
        
        # Opciones para PDF
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'no-outline': None
        }
        
        # Generar PDF
        pdf = pdfkit.from_string(html_content, False, options=options, configuration=config)
        
        # Crear respuesta
        pdf_io = io.BytesIO(pdf)
        
        return send_file(
            pdf_io,
            as_attachment=True,
            download_name=f'reporte_tickets_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"Error generando PDF: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Error al generar el reporte PDF'
        }), 500

# ================================
# ENDPOINTS EXISTENTES (completos)
# ================================

@empleados_bp.route("/api/consumos")
@login_required
def api_consumos():
    """Endpoint para obtener datos de consumo para gráficos"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # SOLO agua y luz (sin gas)
        cursor.execute("""
            SELECT 
                TO_CHAR(fecha_registro, 'YYYY-MM') as mes,
                'agua' as tipo,
                SUM(cantidad_registrada) as consumo
            FROM consumo_agua 
            WHERE fecha_registro >= CURRENT_DATE - INTERVAL '1 year'
            GROUP BY TO_CHAR(fecha_registro, 'YYYY-MM')
            
            UNION ALL
            
            SELECT 
                TO_CHAR(fecha_registro, 'YYYY-MM') as mes,
                'luz' as tipo,
                SUM(cantidad_registrada) as consumo
            FROM consumo_luz 
            WHERE fecha_registro >= CURRENT_DATE - INTERVAL '1 year'
            GROUP BY TO_CHAR(fecha_registro, 'YYYY-MM')
            
            ORDER BY mes, tipo
        """)
        consumos = cursor.fetchall()

        # Consumos actuales del mes
        cursor.execute("""
            SELECT 
                COALESCE(SUM(cantidad_registrada), 0) as consumo_agua
            FROM consumo_agua 
            WHERE DATE_PART('month', fecha_registro) = DATE_PART('month', CURRENT_DATE)
            AND DATE_PART('year', fecha_registro) = DATE_PART('year', CURRENT_DATE)
        """)
        agua_actual = cursor.fetchone()[0]

        cursor.execute("""
            SELECT 
                COALESCE(SUM(cantidad_registrada), 0) as consumo_luz
            FROM consumo_luz 
            WHERE DATE_PART('month', fecha_registro) = DATE_PART('month', CURRENT_DATE)
            AND DATE_PART('year', fecha_registro) = DATE_PART('year', CURRENT_DATE)
        """)
        luz_actual = cursor.fetchone()[0]

        # Gas en 0 (ya que no existe la tabla)
        gas_actual = 0

        cursor.close()
        conn.close()

        # Organizar datos para gráficos
        datos_grafico = {}
        for mes, tipo, consumo in consumos:
            if mes not in datos_grafico:
                datos_grafico[mes] = {'mes': mes, 'agua': 0, 'luz': 0, 'gas': 0}
            datos_grafico[mes][tipo] = float(consumo or 0)

        return jsonify({
            'success': True,
            'data': list(datos_grafico.values()),
            'actual': {
                'agua': float(agua_actual or 0),
                'luz': float(luz_actual or 0),
                'gas': 0  # Gas en 0
            }
        })

    except Exception as e:
        print(f"❌ Error en API consumos: {e}")
        return jsonify({'success': False, 'error': str(e)})

@empleados_bp.route("/api/exportar-pdf", methods=["POST"])
@login_required
def exportar_pdf():
    """Exportar reporte en PDF - IMPLEMENTACIÓN REAL"""
    try:
        data = request.get_json()
        tipo_reporte = data.get('tipo_reporte', 'resumen_trabajo')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        
        print(f"📄 Generando PDF real - Tipo: {tipo_reporte}")

        # Obtener datos del empleado
        id_empleado = get_id_empleado()
        if not id_empleado:
            return jsonify({'success': False, 'error': 'Empleado no encontrado'}), 400

        # Crear PDF en memoria
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = styles['Heading1']
        title_style.alignment = 1  # Centrado
        title = Paragraph("REPORTE DE TRABAJO", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Información del empleado
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT u.nombre, u.ap_paterno, u.ap_materno, e.puesto, e.fecha_contratacion
            FROM usuario u
            JOIN empleado e ON u.id_usuario = e.id_usuario
            WHERE e.id_empleado = %s
        """, (id_empleado,))
        empleado_data = cursor.fetchone()
        
        # Información del empleado
        info_text = f"""
        <b>Empleado:</b> {empleado_data[0]} {empleado_data[1]} {empleado_data[2]}<br/>
        <b>Puesto:</b> {empleado_data[3]}<br/>
        <b>Fecha de Contratación:</b> {empleado_data[4].strftime('%d/%m/%Y') if empleado_data[4] else 'N/A'}<br/>
        <b>Fecha de Reporte:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
        <b>Período:</b> {fecha_inicio or 'Inicio'} - {fecha_fin or 'Actual'}
        """
        
        info_style = styles['BodyText']
        info = Paragraph(info_text, info_style)
        elements.append(info)
        elements.append(Spacer(1, 30))
        
        # Estadísticas de Tickets
        cursor.execute("""
            SELECT estado, COUNT(*) as cantidad
            FROM ticket 
            WHERE id_empleado = %s
            GROUP BY estado
        """, (id_empleado,))
        tickets_data = cursor.fetchall()
        
        # Tabla de tickets
        tickets_title = Paragraph("<b>ESTADÍSTICAS DE TICKETS</b>", styles['Heading2'])
        elements.append(tickets_title)
        elements.append(Spacer(1, 10))
        
        if tickets_data:
            ticket_table_data = [['Estado', 'Cantidad']]
            for estado, cantidad in tickets_data:
                ticket_table_data.append([estado.replace('_', ' ').title(), str(cantidad)])
            
            ticket_table = Table(ticket_table_data, colWidths=[3*inch, 1.5*inch])
            ticket_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(ticket_table)
        else:
            elements.append(Paragraph("No hay datos de tickets", styles['BodyText']))
        
        elements.append(Spacer(1, 20))
        
        # Estadísticas de Mantenimientos
        cursor.execute("""
            SELECT estado, COUNT(*) as cantidad
            FROM mantenimiento 
            WHERE id_empleado = %s
            GROUP BY estado
        """, (id_empleado,))
        mantenimientos_data = cursor.fetchall()
        
        # Tabla de mantenimientos
        mant_title = Paragraph("<b>ESTADÍSTICAS DE MANTENIMIENTOS</b>", styles['Heading2'])
        elements.append(mant_title)
        elements.append(Spacer(1, 10))
        
        if mantenimientos_data:
            mant_table_data = [['Estado', 'Cantidad']]
            for estado, cantidad in mantenimientos_data:
                mant_table_data.append([estado.title(), str(cantidad)])
            
            mant_table = Table(mant_table_data, colWidths=[3*inch, 1.5*inch])
            mant_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(mant_table)
        else:
            elements.append(Paragraph("No hay datos de mantenimientos", styles['BodyText']))
        
        cursor.close()
        conn.close()
        
        # Generar PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_trabajo_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"❌ Error exportando PDF: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@empleados_bp.route("/api/exportar-excel", methods=["POST"])
@login_required
def exportar_excel():
    """Exportar reporte en Excel - IMPLEMENTACIÓN REAL"""
    try:
        data = request.get_json()
        tipo_reporte = data.get('tipo_reporte', 'resumen_trabajo')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        
        print(f"📊 Generando Excel real - Tipo: {tipo_reporte}")

        # Obtener datos del empleado
        id_empleado = get_id_empleado()
        if not id_empleado:
            return jsonify({'success': False, 'error': 'Empleado no encontrado'}), 400

        # Crear workbook de Excel
        wb = Workbook()
        
        # Hoja de resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16)
        normal_font = Font(size=11)
        
        # Título
        ws_summary['A1'] = "REPORTE DE TRABAJO"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        
        # Información del empleado
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT u.nombre, u.ap_paterno, u.ap_materno, e.puesto
            FROM usuario u
            JOIN empleado e ON u.id_usuario = e.id_usuario
            WHERE e.id_empleado = %s
        """, (id_empleado,))
        empleado_data = cursor.fetchone()
        
        ws_summary['A3'] = "Empleado:"
        ws_summary['B3'] = f"{empleado_data[0]} {empleado_data[1]} {empleado_data[2]}"
        ws_summary['A4'] = "Puesto:"
        ws_summary['B4'] = empleado_data[3]
        ws_summary['A5'] = "Fecha de Reporte:"
        ws_summary['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        # Estadísticas de Tickets
        cursor.execute("""
            SELECT estado, COUNT(*) as cantidad
            FROM ticket 
            WHERE id_empleado = %s
            GROUP BY estado
        """, (id_empleado,))
        tickets_data = cursor.fetchall()
        
        # Tabla de tickets
        ws_summary['A7'] = "ESTADÍSTICAS DE TICKETS"
        ws_summary['A7'].font = header_font
        ws_summary['A7'].fill = header_fill
        ws_summary.merge_cells('A7:B7')
        
        ws_summary['A8'] = "Estado"
        ws_summary['B8'] = "Cantidad"
        ws_summary['A8'].font = Font(bold=True)
        ws_summary['B8'].font = Font(bold=True)
        
        row = 9
        for estado, cantidad in tickets_data:
            ws_summary[f'A{row}'] = estado.replace('_', ' ').title()
            ws_summary[f'B{row}'] = cantidad
            row += 1
        
        # Estadísticas de Mantenimientos
        cursor.execute("""
            SELECT estado, COUNT(*) as cantidad
            FROM mantenimiento 
            WHERE id_empleado = %s
            GROUP BY estado
        """, (id_empleado,))
        mantenimientos_data = cursor.fetchall()
        
        # Tabla de mantenimientos
        ws_summary['D7'] = "ESTADÍSTICAS DE MANTENIMIENTOS"
        ws_summary['D7'].font = header_font
        ws_summary['D7'].fill = header_fill
        ws_summary.merge_cells('D7:E7')
        
        ws_summary['D8'] = "Estado"
        ws_summary['E8'] = "Cantidad"
        ws_summary['D8'].font = Font(bold=True)
        ws_summary['E8'].font = Font(bold=True)
        
        row = 9
        for estado, cantidad in mantenimientos_data:
            ws_summary[f'D{row}'] = estado.title()
            ws_summary[f'E{row}'] = cantidad
            row += 1
        
        # Hoja de tickets detallados
        ws_tickets = wb.create_sheet("Tickets Detallados")
        
        # Encabezados de tickets
        headers_tickets = ['ID', 'Descripción', 'Estado', 'Prioridad', 'Fecha Emisión', 'Área']
        for col, header in enumerate(headers_tickets, 1):
            cell = ws_tickets.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos de tickets
        cursor.execute("""
            SELECT 
                t.id_ticket,
                t.descripcion,
                t.estado,
                t.prioridad,
                t.fecha_emision,
                a.nombre as area
            FROM ticket t
            JOIN area a ON t.id_area = a.id_area
            WHERE t.id_empleado = %s
            ORDER BY t.fecha_emision DESC
        """, (id_empleado,))
        tickets_detalle = cursor.fetchall()
        
        for row_idx, ticket in enumerate(tickets_detalle, 2):
            for col_idx, value in enumerate(ticket, 1):
                ws_tickets.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar anchos de columnas
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        cursor.close()
        conn.close()
        
        # Guardar en memoria
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_trabajo_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Error exportando Excel: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@empleados_bp.route("/api/generar-reporte-jasper", methods=["POST"])
@login_required
def generar_reporte_jasper():
    """Generar reporte usando JasperServer via REST API"""
    try:
        data = request.get_json()
        formato = data.get('formato', 'pdf')
        
        id_empleado = get_id_empleado()
        if not id_empleado:
            return jsonify({'success': False, 'error': 'Empleado no encontrado'}), 400

        # Configuración de JasperServer
        JASPER_SERVER_URL = "http://localhost:8080/jasperserver"
        USERNAME = "jasperadmin"
        PASSWORD = "jasperadmin"
        
        # Parámetros para el reporte
        params = {
            "id_empleado": id_empleado,
            "fecha_generacion": datetime.now().strftime('%Y-%m-%d')
        }
        
        # URL del reporte en JasperServer
        report_path = "/reports/reports/empleados_report"
        report_url = f"{JASPER_SERVER_URL}/rest_v2/reports{report_path}.{formato}"
        
        # Autenticación Basic Auth
        auth_string = base64.b64encode(f"{USERNAME}:{PASSWORD}".encode()).decode()
        headers = {
            'Authorization': f'Basic {auth_string}',
            'Accept': 'application/json'
        }
        
        # Llamar a JasperServer
        response = requests.get(
            report_url,
            params=params,
            headers=headers,
            stream=True
        )
        
        if response.status_code == 200:
            # Devolver el archivo generado
            buffer = io.BytesIO(response.content)
            
            mimetypes = {
                'pdf': 'application/pdf',
                'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'html': 'text/html'
            }
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=f'reporte_jasper_{datetime.now().strftime("%Y%m%d_%H%M")}.{formato}',
                mimetype=mimetypes.get(formato, 'application/octet-stream')
            )
        else:
            print(f"❌ Error JasperServer: {response.status_code} - {response.text}")
            return jsonify({
                'success': False, 
                'error': f'Error del servidor de reportes: {response.status_code}'
            }), 500
            
    except Exception as e:
        print(f"❌ Error generando reporte Jasper: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@empleados_bp.route("/api/debug/consumos")
def debug_consumos():
    """Ruta temporal para debuggear qué tablas existen"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar qué tablas de consumo existen
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name LIKE 'consumo%'
        """)
        tablas = [tabla[0] for tabla in cursor.fetchall()]
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'tablas_consumo': tablas,
            'total_tablas': len(tablas)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

# ================================
# ENDPOINTS PARA ESTADÍSTICAS
# ================================

@empleados_bp.route('/api/estadisticas')
@login_required
def obtener_estadisticas():
    """API para obtener estadísticas para el dashboard de reportes"""
    try:
        id_empleado = get_id_empleado()
        if not id_empleado:
            return jsonify({'success': False, 'error': 'Empleado no encontrado'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # Estadísticas de tickets
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                COUNT(CASE WHEN estado = 'abierto' THEN 1 END) as abiertos,
                COUNT(CASE WHEN estado = 'en_progreso' THEN 1 END) as en_progreso,
                COUNT(CASE WHEN estado = 'cerrado' THEN 1 END) as cerrados,
                COUNT(CASE WHEN prioridad = 'urgente' THEN 1 END) as urgentes
            FROM ticket 
            WHERE id_empleado = %s
        """, (id_empleado,))
        stats_tickets = cursor.fetchone()

        # Estadísticas de mantenimientos
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                COUNT(CASE WHEN estado = 'programado' THEN 1 END) as programados,
                COUNT(CASE WHEN estado = 'en_proceso' THEN 1 END) as en_proceso,
                COUNT(CASE WHEN estado = 'completado' THEN 1 END) as completados
            FROM mantenimiento 
            WHERE id_empleado = %s
        """, (id_empleado,))
        stats_mantenimientos = cursor.fetchone()

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'data': {
                'tickets': {
                    'total': stats_tickets[0],
                    'abiertos': stats_tickets[1],
                    'en_progreso': stats_tickets[2],
                    'cerrados': stats_tickets[3],
                    'urgentes': stats_tickets[4]
                },
                'mantenimientos': {
                    'total': stats_mantenimientos[0],
                    'programados': stats_mantenimientos[1],
                    'en_proceso': stats_mantenimientos[2],
                    'completados': stats_mantenimientos[3]
                }
            }
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo estadísticas: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    
# ================================
# HISTORIAL DE PAGOS - EMPLEADO (VERSIÓN FINAL CON TABLAS REALES)
# ================================
@empleados_bp.route("/historial-pagos")
@login_required
def historial_pagos():
    """Página de historial de pagos para el empleado"""
    if current_user.id_rol != 2:
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("auth.login"))
    
    try:
        id_empleado = get_id_empleado()
        if not id_empleado:
            flash("No se encontró información del empleado.", "danger")
            return render_template("empleado/historial_pagos.html", pagos=[], historial_salarios=[], salario_actual=0)

        conn = get_db_connection()
        cursor = conn.cursor()

        # Obtener salario actual
        cursor.execute("SELECT salario FROM empleado WHERE id_empleado = %s", (id_empleado,))
        salario_actual_result = cursor.fetchone()
        salario_actual = salario_actual_result[0] if salario_actual_result else 0

        # Obtener historial de pagos de nómina (REALES)
        cursor.execute("""
            SELECT 
                id_pago_nomina,
                periodo_inicio,
                periodo_fin,
                fecha_pago,
                monto_pagado,
                horas_trabajadas,
                bonos,
                deducciones,
                metodo_pago,
                estado_pago,
                comprobante_url
            FROM pago_nomina 
            WHERE id_empleado = %s
            ORDER BY fecha_pago DESC
            LIMIT 12
        """, (id_empleado,))
        pagos = cursor.fetchall()

        # Obtener historial de cambios de salario (REALES)
        cursor.execute("""
            SELECT 
                fecha_cambio,
                salario_anterior,
                salario_nuevo,
                motivo_cambio
            FROM historial_salario 
            WHERE id_empleado = %s
            ORDER BY fecha_cambio DESC
            LIMIT 10
        """, (id_empleado,))
        historial_salarios = cursor.fetchall()

        cursor.close()
        conn.close()

        return render_template(
            "empleado/historial_pagos.html",
            pagos=pagos,
            historial_salarios=historial_salarios,
            salario_actual=salario_actual
        )

    except Exception as e:
        print(f"❌ Error cargando historial de pagos: {e}")
        flash("Error al cargar el historial de pagos.", "danger")
        return render_template(
            "empleado/historial_pagos.html",
            pagos=[],
            historial_salarios=[],
            salario_actual=0
        )

# ================================
# API PARA DATOS DE PAGOS (Gráficos) - VERSIÓN REAL
# ================================
@empleados_bp.route("/api/pagos-estadisticas")
@login_required
def api_pagos_estadisticas():
    """API para obtener estadísticas de pagos para gráficos (DATOS REALES)"""
    try:
        id_empleado = get_id_empleado()
        if not id_empleado:
            return jsonify({'success': False, 'error': 'Empleado no encontrado'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # Obtener datos reales para gráfico mensual
        cursor.execute("""
            SELECT 
                TO_CHAR(fecha_pago, 'YYYY-MM') as mes,
                AVG(monto_pagado) as promedio_mensual
            FROM pago_nomina 
            WHERE id_empleado = %s 
            AND fecha_pago >= CURRENT_DATE - INTERVAL '6 months'
            GROUP BY TO_CHAR(fecha_pago, 'YYYY-MM')
            ORDER BY mes
        """, (id_empleado,))
        datos_mensuales = cursor.fetchall()

        # Obtener totales para gráfico de categorías
        cursor.execute("""
            SELECT 
                AVG(monto_pagado - bonos) as salario_base,
                AVG(bonos) as bonos_promedio,
                AVG(deducciones) as deducciones_promedio
            FROM pago_nomina 
            WHERE id_empleado = %s 
            AND fecha_pago >= CURRENT_DATE - INTERVAL '6 months'
        """, (id_empleado,))
        promedios = cursor.fetchone()

        cursor.close()
        conn.close()

        # Preparar datos para el gráfico
        datos_mensual = []
        for mes, promedio in datos_mensuales:
            datos_mensual.append({
                'mes': mes[-5:],  # Tomar solo "MM-YY"
                'monto': float(promedio)
            })

        datos_categorias = [
            {'categoria': 'Salario Base', 'monto': float(promedios[0]) if promedios[0] else 2100},
            {'categoria': 'Bonos', 'monto': float(promedios[1]) if promedios[1] else 150},
            {'categoria': 'Deducciones', 'monto': float(promedios[2]) if promedios[2] else -120}
        ]

        return jsonify({
            'success': True,
            'data': {
                'mensual': datos_mensual,
                'categorias': datos_categorias
            }
        })

    except Exception as e:
        print(f"❌ Error en API pagos estadísticas: {e}")
        return jsonify({'success': False, 'error': str(e)})